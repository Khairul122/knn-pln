"""
============================================================
  KNN PREDIKSI RISIKO GANGGUAN JARINGAN LISTRIK
  PLN ULP Lhokseumawe — 2024
  ============================================================
  Pipeline lengkap:
    1. Load & Preprocessing
    2. Normalisasi Min-Max
    3. Split Train / Test (80:20)
    4. Cross-Validation → K Terbaik
    5. Euclidean Distance (manual + sklearn)
    6. Training & Prediksi
    7. Evaluasi (Confusion Matrix, Accuracy, Precision, Recall, F1)
    8. Export Hasil ke Excel
    9. Save Model (deploy)
============================================================
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import joblib
import warnings
warnings.filterwarnings('ignore')

from sklearn.neighbors        import KNeighborsClassifier
from sklearn.preprocessing    import LabelEncoder, MinMaxScaler
from sklearn.model_selection  import train_test_split, cross_val_score, StratifiedKFold
from sklearn.metrics          import (confusion_matrix, classification_report,
                                      accuracy_score, precision_score,
                                      recall_score, f1_score)

# ── palet warna konsisten ─────────────────────────────────
CLR = {"Rendah": "#2ECC71", "Sedang": "#F39C12", "Tinggi": "#E74C3C"}
PLT_STYLE = {"figure.facecolor": "white", "axes.facecolor": "white",
             "axes.grid": True, "grid.alpha": 0.3}
plt.rcParams.update(PLT_STYLE)
plt.rcParams['font.family'] = 'DejaVu Sans'

OUTPUT_DIR = "c:\\Users\\ASUS\\Downloads\\perhitungan data"

# ╔══════════════════════════════════════════════════════════╗
# ║  1. LOAD DATA                                            ║
# ╚══════════════════════════════════════════════════════════╝
print("=" * 60)
print("  PLN ULP LHOKSEUMAWE — PREDIKSI RISIKO GANGGUAN (KNN)")
print("=" * 60)

df = pd.read_excel("PEMELIHARAAN_PLN_2024_CLEANED.xlsx")
print(f"\n[1] Dataset dimuat  →  {df.shape[0]} baris × {df.shape[1]} kolom")

# Hitung RPN & Label
def severity(row):
    for col in ['PERGANTIAN_FCO', 'PERGANTIAN FCO', 'PERBAIKAN GROUNDING TRAFO']:
        if col in row.index and row.get(col, 0) > 0: return 9
    if row.get('PENYEIMBANGAN BEBAN GARDU', 0) > 0: return 6
    if row.get('PENGUKURAN', 0) > 0: return 4
    return 1

def occurrence(row):
    total = row.get('TIER1_TEMUAN', 0) + row.get('TIER1 TEMUAN', 0) + row.get('TIER2_TEMUAN', 0) + row.get('TIER2 TEMUAN', 0)
    return 1 if total == 0 else (4 if total <= 10 else (7 if total <= 20 else 9))

def detection(row):
    tier1 = row.get('TIER1_INPEKSI', 0) + row.get('TIER1 INPEKSI', 0)
    tier2 = row.get('TIER2_INPEKSI', 0) + row.get('TIER2 INPEKSI', 0)
    return 2 if tier1 > 0 and tier2 > 0 else (5 if tier1 > 0 or tier2 > 0 else 9)

df['S'] = df.apply(severity, axis=1)
df['O'] = df.apply(occurrence, axis=1)
df['D'] = df.apply(detection, axis=1)
df['RPN'] = df['S'] * df['O'] * df['D']
df['LABEL'] = df['RPN'].apply(lambda x: 'Rendah' if x <= 9 else ('Sedang' if x <= 99 else 'Tinggi'))

FEATURES = [
    "TIER1_INPEKSI", "TIER1_TEMUAN",
    "TIER2_INPEKSI", "TIER2_TEMUAN",
    "PENGUKURAN", "PERGANTIAN_FCO",
    "PENYEIMBANGAN BEBAN GARDU",
    "PERBAIKAN GROUNDING TRAFO",
    "PENGHALANG_PANJAT"
]
TARGET = "LABEL"

print(f"    Fitur input  : {len(FEATURES)} variabel")
print(f"    Target label : {TARGET}")
print(f"\n    Distribusi label:")
for lbl, cnt in df[TARGET].value_counts().items():
    pct = cnt / len(df) * 100
    print(f"      {lbl:8s}: {cnt:4d} ({pct:.1f}%)")

# ╔══════════════════════════════════════════════════════════╗
# ║  2. PREPROCESSING                                        ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[2] Preprocessing …")

df[FEATURES] = df[FEATURES].fillna(0)
X_raw = df[FEATURES].values.astype(float)
y_str = df[TARGET].values

le = LabelEncoder()
y = le.fit_transform(y_str)            # Rendah=0, Sedang=1, Tinggi=2
CLASS_NAMES = le.classes_             # ['Rendah', 'Sedang', 'Tinggi']
print(f"    Encoding label : {dict(zip(CLASS_NAMES, le.transform(CLASS_NAMES)))}")

# ── Normalisasi Min-Max ───────────────────────────────────
scaler = MinMaxScaler()
X = scaler.fit_transform(X_raw)

print("    Min-Max Scaling :")
for i, feat in enumerate(FEATURES):
    mn, mx = scaler.data_min_[i], scaler.data_max_[i]
    print(f"      {feat:<35s}  min={mn:.2f}  max={mx:.2f}")

# ╔══════════════════════════════════════════════════════════╗
# ║  3. SPLIT TRAIN / TEST  80 : 20                          ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[3] Train / Test Split  (80:20, stratified) …")
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.20, random_state=42, stratify=y
)
print(f"    Training set : {X_train.shape[0]} data")
print(f"    Testing  set : {X_test.shape[0]} data")

# simpan indeks asli untuk laporan
idx_train, idx_test = train_test_split(
    np.arange(len(df)), test_size=0.20, random_state=42, stratify=y
)

# ╔══════════════════════════════════════════════════════════╗
# ║  4. CROSS-VALIDATION → K TERBAIK                        ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[4] Cross-Validation (Stratified 10-Fold) untuk mencari K terbaik …")

K_RANGE = range(1, 22, 2)   # K ganjil: 1,3,5,…,21
cv = StratifiedKFold(n_splits=10, shuffle=True, random_state=42)

cv_results = {}
for k in K_RANGE:
    knn = KNeighborsClassifier(n_neighbors=k, metric='euclidean')
    scores = cross_val_score(knn, X_train, y_train, cv=cv, scoring='accuracy')
    cv_results[k] = scores

cv_df = pd.DataFrame(cv_results).T
cv_df.index.name = "K"
cv_df.columns = [f"Fold_{i+1}" for i in range(10)]
cv_df["Mean_Accuracy"] = cv_df.mean(axis=1)
cv_df["Std"] = cv_df.std(axis=1)

print(f"\n    {'K':>4} │ {'Mean Accuracy':>14} │ {'Std':>8}")
print("    " + "─" * 32)
for k, row in cv_df.iterrows():
    marker = " ◄ TERBAIK" if row["Mean_Accuracy"] == cv_df["Mean_Accuracy"].max() else ""
    print(f"    {k:>4} │ {row['Mean_Accuracy']:>13.4f} │ {row['Std']:>8.4f}{marker}")

K_BEST = int(cv_df["Mean_Accuracy"].idxmax())
print(f"\n    ✓  K terbaik = {K_BEST}  "
      f"(accuracy = {cv_df.loc[K_BEST,'Mean_Accuracy']:.4f} ± "
      f"{cv_df.loc[K_BEST,'Std']:.4f})")

# ── Plot CV ───────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 4))
ks = list(cv_results.keys())
means = cv_df["Mean_Accuracy"].values
stds  = cv_df["Std"].values
ax.fill_between(ks, means - stds, means + stds, alpha=0.15, color="#2980B9")
ax.plot(ks, means, "o-", color="#2980B9", lw=2, ms=6, label="Mean Accuracy (CV)")
ax.axvline(K_BEST, ls="--", color="#E74C3C", lw=1.5,
           label=f"K terbaik = {K_BEST}  ({means[ks.index(K_BEST)]:.4f})")
ax.scatter([K_BEST], [cv_df.loc[K_BEST,"Mean_Accuracy"]],
           s=120, color="#E74C3C", zorder=5)
ax.set_xlabel("Nilai K", fontsize=11)
ax.set_ylabel("Akurasi (10-Fold CV)", fontsize=11)
ax.set_title("Cross-Validation: Pencarian K Terbaik", fontsize=13, fontweight="bold")
ax.set_xticks(list(K_RANGE))
ax.legend(fontsize=10)
plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/plot_cv_k.png", dpi=150)
plt.close()
print("    → Grafik CV disimpan: plot_cv_k.png")

# ╔══════════════════════════════════════════════════════════╗
# ║  5. EUCLIDEAN DISTANCE — ILUSTRASI MANUAL                ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[5] Demonstrasi perhitungan Euclidean Distance …")
print("    (3 data uji pertama vs seluruh data latih)")

DEMO_INDICES = [0, 1, 2]
demo_results = []

for di in DEMO_INDICES:
    x_q = X_test[di]
    dists = np.sqrt(np.sum((X_train - x_q) ** 2, axis=1))
    top_k_idx = np.argsort(dists)[:K_BEST]

    orig_q = df.iloc[idx_test[di]]
    print(f"\n    ── Data Uji [{di}]  {orig_q['PENYULANG']} / {orig_q['BULAN']} "
          f"(label asli: {orig_q[TARGET]}) ──")
    print(f"    {'Rank':<5} {'Penyulang':<25} {'Bulan':<12} "
          f"{'Jarak':>8}  {'Label':<8}")
    print("    " + "─" * 65)
    for rank, tidx in enumerate(top_k_idx, 1):
        orig_t = df.iloc[idx_train[tidx]]
        print(f"    {rank:<5} {str(orig_t['PENYULANG']):<25} "
              f"{str(orig_t['BULAN']):<12} "
              f"{dists[tidx]:>8.4f}  {orig_t[TARGET]}")

    # voting
    votes = y_train[top_k_idx]
    vote_count = {CLASS_NAMES[c]: int(np.sum(votes == c))
                  for c in range(len(CLASS_NAMES))}
    pred_label = CLASS_NAMES[np.bincount(votes).argmax()]
    confidence = max(vote_count.values()) / K_BEST * 100
    print(f"\n    Voting  : {vote_count}")
    print(f"    Prediksi: {pred_label}   Confidence: {confidence:.1f}%")

    demo_results.append({
        "Penyulang": orig_q["PENYULANG"], "Bulan": orig_q["BULAN"],
        "Label_Asli": orig_q[TARGET], "Prediksi": pred_label,
        "Confidence (%)": round(confidence, 1), **vote_count
    })

# ╔══════════════════════════════════════════════════════════╗
# ║  6. TRAINING MODEL FINAL & PREDIKSI                      ║
# ╚══════════════════════════════════════════════════════════╝
print(f"\n[6] Training model KNN (K={K_BEST}, metric=euclidean) …")

knn_final = KNeighborsClassifier(n_neighbors=K_BEST, metric='euclidean')
knn_final.fit(X_train, y_train)

y_pred = knn_final.predict(X_test)
y_pred_proba = knn_final.predict_proba(X_test)

print(f"    ✓  Model terlatih  |  {X_train.shape[0]} data training")

# ╔══════════════════════════════════════════════════════════╗
# ║  7. EVALUASI                                             ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[7] Evaluasi Model …")

acc = accuracy_score(y_test, y_pred)
prec_macro = precision_score(y_test, y_pred, average='macro', zero_division=0)
rec_macro  = recall_score   (y_test, y_pred, average='macro', zero_division=0)
f1_macro   = f1_score       (y_test, y_pred, average='macro', zero_division=0)

prec_w = precision_score(y_test, y_pred, average='weighted', zero_division=0)
rec_w  = recall_score   (y_test, y_pred, average='weighted', zero_division=0)
f1_w   = f1_score       (y_test, y_pred, average='weighted', zero_division=0)

print(f"\n    {'Metrik':<30} {'Macro':>10} {'Weighted':>10}")
print("    " + "─" * 52)
print(f"    {'Accuracy':<30} {acc:>10.4f}")
print(f"    {'Precision':<30} {prec_macro:>10.4f} {prec_w:>10.4f}")
print(f"    {'Recall':<30} {rec_macro:>10.4f} {rec_w:>10.4f}")
print(f"    {'F1-Score':<30} {f1_macro:>10.4f} {f1_w:>10.4f}")

print("\n    Laporan Per Kelas:")
print(classification_report(y_test, y_pred,
                             target_names=CLASS_NAMES,
                             zero_division=0))

# ── Confusion Matrix ──────────────────────────────────────
cm = confusion_matrix(y_test, y_pred)
fig, axes = plt.subplots(1, 2, figsize=(13, 5))

# raw counts
sns.heatmap(cm, annot=True, fmt="d", ax=axes[0],
            xticklabels=CLASS_NAMES, yticklabels=CLASS_NAMES,
            cmap="Blues", linewidths=0.5, linecolor="white",
            annot_kws={"size": 14, "weight": "bold"})
axes[0].set_xlabel("Prediksi", fontsize=11)
axes[0].set_ylabel("Aktual", fontsize=11)
axes[0].set_title(f"Confusion Matrix (count)\nK={K_BEST} | Acc={acc:.4f}",
                  fontsize=12, fontweight="bold")

# normalized
cm_norm = cm.astype(float) / cm.sum(axis=1, keepdims=True)
sns.heatmap(cm_norm, annot=True, fmt=".2f", ax=axes[1],
            xticklabels=CLASS_NAMES, yticklabels=CLASS_NAMES,
            cmap="YlOrRd", vmin=0, vmax=1, linewidths=0.5, linecolor="white",
            annot_kws={"size": 13})
axes[1].set_xlabel("Prediksi", fontsize=11)
axes[1].set_ylabel("Aktual", fontsize=11)
axes[1].set_title("Confusion Matrix (normalized)", fontsize=12, fontweight="bold")

plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/plot_confusion_matrix.png", dpi=150)
plt.close()
print("    → Confusion matrix disimpan: plot_confusion_matrix.png")

# ── Per-Class bar chart ───────────────────────────────────
per_class = precision_score(y_test, y_pred, average=None, zero_division=0)
per_class_r = recall_score (y_test, y_pred, average=None, zero_division=0)
per_class_f = f1_score     (y_test, y_pred, average=None, zero_division=0)

fig, ax = plt.subplots(figsize=(9, 4))
x = np.arange(len(CLASS_NAMES))
w = 0.25
bars1 = ax.bar(x - w, per_class,   w, label="Precision", color="#3498DB", alpha=0.85)
bars2 = ax.bar(x,     per_class_r, w, label="Recall",    color="#2ECC71", alpha=0.85)
bars3 = ax.bar(x + w, per_class_f, w, label="F1-Score",  color="#E67E22", alpha=0.85)
for bars in [bars1, bars2, bars3]:
    for b in bars:
        ax.text(b.get_x() + b.get_width()/2, b.get_height() + 0.01,
                f"{b.get_height():.2f}", ha='center', va='bottom', fontsize=9)
ax.axhline(acc, ls="--", color="#E74C3C", lw=1.3, label=f"Accuracy={acc:.4f}")
ax.set_xticks(x); ax.set_xticklabels(CLASS_NAMES, fontsize=11)
ax.set_ylim(0, 1.12); ax.set_ylabel("Score", fontsize=11)
ax.set_title(f"Evaluasi Per Kelas  (K={K_BEST})", fontsize=13, fontweight="bold")
ax.legend(fontsize=10); plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/plot_metrics_per_class.png", dpi=150)
plt.close()
print("    → Grafik metrik per kelas disimpan: plot_metrics_per_class.png")

# ── Distribusi prediksi vs aktual ─────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(10, 4))
for ax, (labels, title) in zip(axes, [
        (y_test,  "Label Aktual  (Test Set)"),
        (y_pred,  "Label Prediksi (KNN)")]):
    counts = {c: int(np.sum(labels == i)) for i, c in enumerate(CLASS_NAMES)}
    bars = ax.bar(counts.keys(), counts.values(),
                  color=[CLR[c] for c in counts], edgecolor="white",
                  linewidth=0.8, alpha=0.9)
    for b, v in zip(bars, counts.values()):
        ax.text(b.get_x()+b.get_width()/2, b.get_height()+0.4,
                str(v), ha='center', fontsize=12, fontweight='bold')
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.set_ylabel("Jumlah Data"); ax.set_ylim(0, max(counts.values())+8)
plt.tight_layout()
plt.savefig(f"{OUTPUT_DIR}/plot_distribusi_prediksi.png", dpi=150)
plt.close()
print("    → Grafik distribusi prediksi disimpan: plot_distribusi_prediksi.png")

# ╔══════════════════════════════════════════════════════════╗
# ║  8. EXPORT HASIL LENGKAP KE EXCEL                        ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[8] Menyimpan hasil ke Excel …")

from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers as xlnumbers)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Siapkan DataFrame hasil ──
df_train_out = df.iloc[idx_train].copy().reset_index(drop=True)
df_train_out["Split"] = "Training"
df_train_out["Prediksi"] = CLASS_NAMES[knn_final.predict(X_train)]

df_test_out = df.iloc[idx_test].copy().reset_index(drop=True)
df_test_out["Split"] = "Testing"
df_test_out["Prediksi"] = CLASS_NAMES[y_pred]

# Confidence (probability maks)
conf_train = knn_final.predict_proba(X_train).max(axis=1) * 100
conf_test  = y_pred_proba.max(axis=1) * 100
df_train_out["Confidence (%)"] = np.round(conf_train, 1)
df_test_out ["Confidence (%)"] = np.round(conf_test,  1)
df_test_out ["Benar/Salah"]    = np.where(y_pred == y_test, "✓ Benar", "✗ Salah")

df_cv_export = cv_df.reset_index()

# ── Metrik ringkasan ──
def eval_row(y_t, y_p, split):
    return {
        "Split": split,
        "Total Data": len(y_t),
        "Accuracy": round(accuracy_score(y_t, y_p), 4),
        "Precision (Macro)": round(precision_score(y_t, y_p, average='macro', zero_division=0), 4),
        "Recall (Macro)": round(recall_score(y_t, y_p, average='macro', zero_division=0), 4),
        "F1-Score (Macro)": round(f1_score(y_t, y_p, average='macro', zero_division=0), 4),
        "Precision (Weighted)": round(precision_score(y_t, y_p, average='weighted', zero_division=0), 4),
        "Recall (Weighted)": round(recall_score(y_t, y_p, average='weighted', zero_division=0), 4),
        "F1-Score (Weighted)": round(f1_score(y_t, y_p, average='weighted', zero_division=0), 4),
    }

df_metrics = pd.DataFrame([
    eval_row(y_train, knn_final.predict(X_train), "Training"),
    eval_row(y_test,  y_pred,                     "Testing"),
])

# confusion matrix → df
cm_df = pd.DataFrame(cm, index=[f"Aktual_{c}" for c in CLASS_NAMES],
                     columns=[f"Pred_{c}" for c in CLASS_NAMES])
cm_df.index.name = "Aktual \\ Prediksi"

# scaler params
scaler_df = pd.DataFrame({"Fitur": FEATURES,
                           "Min": scaler.data_min_,
                           "Max": scaler.data_max_})

# cross val export
cv_export = cv_df.reset_index().rename(columns={"index":"K"})

XLSX_PATH = f"{OUTPUT_DIR}/KNN_PLN_HASIL.xlsx"

with pd.ExcelWriter(XLSX_PATH, engine='openpyxl') as writer:
    df_train_out.to_excel(writer, sheet_name="Data Training",  index=False)
    df_test_out .to_excel(writer, sheet_name="Data Testing",   index=False)
    cv_export   .to_excel(writer, sheet_name="Cross Validation", index=False)
    df_metrics  .to_excel(writer, sheet_name="Metrik Evaluasi", index=False)
    cm_df       .to_excel(writer, sheet_name="Confusion Matrix", index=True)
    scaler_df   .to_excel(writer, sheet_name="Normalisasi",     index=False)
    pd.DataFrame(demo_results).to_excel(
                               writer, sheet_name="Demo Euclidean", index=False)

# ── Styling Excel ──
BLUE_H  = "1F4E79"; LITE_B = "D6E4F0"; AMBER  = "FFF2CC"
GREEN_H = "1E4620"; LITE_G = "E2EFDA"; RED_H  = "641E16"; LITE_R = "FADBD8"
thin = Side(style="thin", color="BFBFBF")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, header_color, light_color, highlight_cols=None):
    hf  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    nf  = Font(name="Calibri", size=10)
    hfl = PatternFill("solid", start_color=header_color, end_color=header_color)
    alt = PatternFill("solid", start_color=light_color,  end_color=light_color)
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = brd
            cell.font   = hf if row_idx == 1 else nf
            if row_idx == 1:
                cell.fill = hfl
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            elif row_idx % 2 == 0:
                cell.fill = alt
            if highlight_cols and cell.column_letter in highlight_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 28)

wb = load_workbook(XLSX_PATH)
style_sheet(wb["Data Training"],   BLUE_H,  LITE_B)
style_sheet(wb["Data Testing"],    BLUE_H,  LITE_B)
style_sheet(wb["Cross Validation"],GREEN_H, LITE_G)
style_sheet(wb["Metrik Evaluasi"], BLUE_H,  LITE_B)
style_sheet(wb["Confusion Matrix"],RED_H,   LITE_R)
style_sheet(wb["Normalisasi"],     BLUE_H,  LITE_B)
style_sheet(wb["Demo Euclidean"],  BLUE_H,  LITE_B)

# Warnai sel Benar/Salah
ws_test = wb["Data Testing"]
for row in ws_test.iter_rows(min_row=2):
    for cell in row:
        if cell.value == "✓ Benar":
            cell.fill = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
            cell.font = Font(color="1E8449", bold=True, name="Calibri", size=10)
        elif cell.value == "✗ Salah":
            cell.fill = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
            cell.font = Font(color="C0392B", bold=True, name="Calibri", size=10)

# Sisipkan gambar CV ke sheet CV
ws_cv = wb["Cross Validation"]
img_cv = XLImage(f"{OUTPUT_DIR}/plot_cv_k.png")
img_cv.width, img_cv.height = 520, 240
ws_cv.add_image(img_cv, "N2")

# Sisipkan confusion matrix ke sheet Metrik
ws_met = wb["Metrik Evaluasi"]
img_cm = XLImage(f"{OUTPUT_DIR}/plot_confusion_matrix.png")
img_cm.width, img_cm.height = 600, 240
ws_met.add_image(img_cm, "L2")

wb.save(XLSX_PATH)
print(f"    ✓  Excel disimpan: KNN_PLN_HASIL.xlsx")

# ╔══════════════════════════════════════════════════════════╗
# ║  9. SAVE MODEL (DEPLOY)                                  ║
# ╚══════════════════════════════════════════════════════════╝
print("\n[9] Menyimpan model untuk deployment …")

MODEL_PKG = {
    "model"        : knn_final,
    "scaler"       : scaler,
    "label_encoder": le,
    "features"     : FEATURES,
    "k_best"       : K_BEST,
    "class_names"  : list(CLASS_NAMES),
    "accuracy"     : round(acc, 4),
    "metadata": {
        "dataset": "PEMELIHARAAN_PLN_2024",
        "n_train": len(X_train),
        "n_test" : len(X_test),
    }
}

joblib.dump(MODEL_PKG, f"{OUTPUT_DIR}/model_knn_pln.pkl")
print(f"    ✓  Model disimpan: model_knn_pln.pkl")

# ── Fungsi prediksi siap pakai ──
def predict_risiko(data: dict) -> dict:
    """
    Prediksi risiko gangguan satu data baru.

    Parameters
    ----------
    data : dict  —  kunci = nama fitur (9 variabel pemeliharaan)

    Returns
    -------
    dict  —  label, confidence, probabilities
    """
    pkg    = joblib.load(f"{OUTPUT_DIR}/model_knn_pln.pkl")
    values = np.array([[data.get(f, 0.0) for f in pkg["features"]]], dtype=float)
    norm   = pkg["scaler"].transform(values)
    pred   = pkg["model"].predict(norm)[0]
    proba  = pkg["model"].predict_proba(norm)[0]
    label  = pkg["label_encoder"].inverse_transform([pred])[0]
    return {
        "label"         : label,
        "confidence (%)" : round(float(proba.max()) * 100, 2),
        "probabilities" : dict(zip(pkg["class_names"],
                                   [round(float(p)*100, 2) for p in proba]))
    }

# ── Demo prediksi data baru ──
print("\n[DEMO DEPLOY] Prediksi satu data baru …")
data_baru = {
    "TIER1_INPEKSI"             : 30,
    "TIER1_TEMUAN"              : 25,
    "TIER2_INPEKSI"             : 0,
    "TIER2_TEMUAN"              : 0,
    "PENGUKURAN"                : 15,
    "PERGANTIAN_FCO"            : 1,
    "PENYEIMBANGAN BEBAN GARDU" : 0,
    "PERBAIKAN GROUNDING TRAFO" : 0,
    "PENGHALANG_PANJAT"         : 8,
}
hasil = predict_risiko(data_baru)
print(f"    Input     : {data_baru}")
print(f"    Prediksi  : {hasil['label']}")
print(f"    Confidence: {hasil['confidence (%)']:.2f}%")
print(f"    Probabilitas: {hasil['probabilities']}")

# ── Ringkasan final ───────────────────────────────────────
print("\n" + "=" * 60)
print("  RINGKASAN HASIL AKHIR")
print("=" * 60)
print(f"  Dataset       : {len(df)} data  ({len(FEATURES)} fitur)")
print(f"  Split         : {len(X_train)} training / {len(X_test)} testing")
print(f"  K terbaik     : {K_BEST}  (10-Fold CV)")
print(f"  Accuracy      : {acc:.4f}  ({acc*100:.2f}%)")
print(f"  Precision (W) : {prec_w:.4f}")
print(f"  Recall    (W) : {rec_w:.4f}")
print(f"  F1-Score  (W) : {f1_w:.4f}")
print(f"  Output        : KNN_PLN_HASIL.xlsx  |  model_knn_pln.pkl")
print("=" * 60)